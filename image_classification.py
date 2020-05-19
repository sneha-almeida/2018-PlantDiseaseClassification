import numpy as np
import cv2
import skimage
from skimage import io
from skimage.io import imread_collection
#from PIL import image
import pandas as pd
import skimage.feature as sk
import numpy as np
import openpyxl
from sklearn import svm
from sklearn.svm import SVC 
from sklearn.model_selection import train_test_split
import openpyxl
from sklearn.preprocessing import normalize
from sklearn.neural_network import MLPClassifier  #Library for BPNN
from sklearn.metrics import classification_report, confusion_matrix
import xlrd
import pickle


images = imread_collection('E:\Tomato___Late_blight\*') #load the images from the data set into python
#io.imshow (images[0])
indexes=[]
filename = "file"
filenum = 1
r = 2


# run this array as many times as there are images which need image processing to be performed on them
for i in range(0,1500):
    #io.imshow (images[i]) #display the images 
    #cv2.imwrite('C:\\Users\Sneha\Desktop\my docs\trial\img1.jpg',images[i])

    image = images[i]
    
    image = cv2.cvtColor(image,cv2.COLOR_BGR2GRAY)#convert the images to gray scale
    
    edges = cv2.Canny(image,100,200) #edge dection algorithm(canny edge detection) is used and a new image which just consists of the edge of original image is stored in edges
    #cv2.imwrite('C:\\Users\Sneha\Desktop\my docs\Elon Musk\image.jpg',edges)
    
    img_final = edges - image #finally on the required pixels are taken and rest of the image is blurred out
    
    #cv2.imwrite('C:\\Users\Sneha\Desktop\my docs\Elon Musk\img1.jpg',img_final)
    print("Values in numpy ")
    print(np.array(img_final)) #print the matrix which makes up the image

    #Extracting the features of an image using greycomatrix and normalizing the resultant matrix
    print("Extracgting features of the image......")
    image = sk.greycomatrix(img_final,[1],[0], 256, symmetric=True, normed=True)

    #image normalization
    #image = cv2.normalize(image, None, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_32F)
    #print(image.shape)
    #image = image.reshape(256,256,1,1)

    #Using the greycoprops function
    #Contrast
    img_contrast = sk.greycoprops(image,'contrast')[0,0]
    a = np.array(img_contrast)
    contrast = np.asscalar(a)
    print("*******************Extracting Feature---------Contrast******************")
    print(img_contrast)
    #Dissimilarity
    img_dissimilarity = sk.greycoprops(image,'dissimilarity')[0,0]
    a = np.array(img_dissimilarity)
    dissimilarity = np.asscalar(a)
    print("*******************Extracting Feature---------dissimilarity******************")
    print(img_dissimilarity)
    #Energy
    img_energy = sk.greycoprops(image,'energy')[0,0]
    a = np.array(img_energy)
    energy = np.asscalar(a)
    print("*******************Extracting Feature---------energy******************")
    print(img_energy)
    #ASM
    img_ASM = sk.greycoprops(image,'ASM')[0,0]
    a = np.array(img_ASM)
    asm = np.asscalar(a)
    print("*******************Extracting Feature---------ASM******************")
    print(img_ASM)
    #Correlation
    img_correlation = sk.greycoprops(image,'correlation')[0,0]
    a = np.array(img_correlation)
    correlation = np.asscalar(a)
    print("*******************Extracting Feature---------correlation******************")
    print(img_correlation)
    #Homogenity
    img_homogeneity = sk.greycoprops(image,'homogeneity')[0,0]
    a = np.array(img_homogeneity)
    homogeneity = np.asscalar(a)
    print("*******************Extracting Feature---------homogeneity******************")
    print(img_homogeneity)
    
    img_matrix = []
    img_matrix = np.array(image) #store it in an array
    print("Values in array........................")
    print(img_matrix)# print that array which is actually the original matrix in a new data container

    
    #creates a file in the directory/folder where the .py file resides and save the matrix to a .txt file
    #np.savetxt('file.txt',(img_matrix),header='Start of the file', footer='End of the file',newline = '\n\n\n')
    #creates a file in the directory/folder where the .py file resides
    
    file = open(filename+str(filenum)+".txt",'w')
    #img_matrix.tofile(file, format = "%s",sep = " ")
    filenum = filenum + 1
    print("************************************Written to file ",i+1,"**********************************************")

    print("Written to file") #acknowledgement printing

    #Finding the dimensions of a matrix
    print("Shape of the greycomatrix",image.shape)
    print("Shape of the contrast matrix",img_contrast.shape)
    print("Shape of the dissimilarity matrix",img_dissimilarity.shape)
    print("Shape of the ASM matrix",img_ASM.shape)
    print("Shape of the correlation matrix",img_correlation.shape)
    print("Shape of the homogeneity matrix",img_homogeneity.shape)

    #calculate the moments for the image
    m = cv2.moments(img_final)
    print("Moment is:",m)


    #calculate the variance of the image
    variance = np.var(img_final)
    print (variance)

    #calculate the mean of the image
    mean = np.mean(img_final)
    print (mean)

    #open the excel file to be updated
               
    wb = openpyxl.load_workbook(filename='E:\plant_classification.xlsx')

    #Gives the number of worksheet in the excel file
    for ws in wb.worksheets:
        print(ws.title)

    #create the data which has to be written to the excel file
    ws = wb.worksheets[0]
    c = [contrast,dissimilarity,asm,correlation,homogeneity,variance,mean,energy]

    #findt the last row where the data is to be appended
    #row = ws.max_row

    #iterates over the excel file to write the data to the excel file
    #here the row will be constant and column will be iterated
    
    #for j in range(len(c)):
        #for l in c.values.tolist():
            #for item in l:
                #ws.cell(row=row, column=j).value = item
  
    
    for j in range(len(c)):
        ws.cell(row=r, column=j+1).value = c[j]
    r = r+1

    #Save the data written to the file
    wb.save("E:\plant_classification_trial.xlsx")
    
    #use fromfile() to reload data
    ##############################Put a comment here##############################


'''
#classifier (SVM)
print("**************************SVM*****************************************")
filename = 'E:\plant_classification.xlsx'
data = pd.read_excel(filename)

x = data.drop('Class', axis=1)
y = data['Class']

#print(x)
#print(y)

train_data,test_data,train_label,test_label =train_test_split(x,y,test_size = 0.1)

#print(train_data)
#print(train_label)
#print(test_data)
#print(test_label)

# training a linear SVM classifier 


print("Creating Classifier..............")
clf = SVC(kernel = 'linear', C = 1)

print("Training..................")
clf.fit(train_data,train_label)
print("Training Complete................")

#testing
predict = clf.predict(test_data)
print(predict)

#Accuracy testing
acc = np.mean(predict==test_label)
print("Accuracy = ",acc*100)

#metrics
print (confusion_matrix(test_label,predict))
print (classification_report(test_label,predict))

#BPNN Classifier
print("****************************************BPNN*******************************")
filename = 'E:\plant_classification.xlsx'
data = pd.read_excel(filename)

z = data.drop('Class', axis=1)
x = z.drop('ASM', axis=1)
y = data['Class']

#print(x)
#print(y)

train_data,test_data,train_label,test_label =train_test_split(x,y,test_size = 0.1)

print("Creating Classifier.........")
mlp = MLPClassifier(hidden_layer_sizes=(100, 100, 100), max_iter=100)
print("training the classifier")
mlp.fit(train_data, train_label)
print("Training Complete")
print(mlp)
#save model to disk
filename = 'finalized_model_bpnn.sav'
pickle.dump(mlp, open(filename, 'wb'))
print("Written to file")

#load the model from the disk
filename = 'finalized_model_bpnn.sav'
loaded_model = pickle.load(open(filename, 'rb'))

predictions = loaded_model.predict(test_data)
print("******Predicted Data********")
print(predictions)

#Accuracy Calculations
accuracy = np.mean(predictions==test_label)
print("Accuracy = ",accuracy*100)  

#metrics
print (confusion_matrix(test_label,predictions))
print (classification_report(test_label,predictions))


#change the names of attributes so that they match the attributes in the excel file 
#suggestions for remedies

#open the excel file and load it in a  data frame

#display the column which suggests a remedy for the predicted disease
disease = "Septoria"

    #loading the data from a csv file into python 
filename = 'E:/remedies.xlsx'
    
    #the data will be stored as a data frame
data = pd.read_excel(filename)
print("*******************The Complete Excel file loaded into python*************************")
print(data)
    
print("*******************Remedies for Septoria*************************")
#iterating over a dataframe in python
#the itertuples is used to iterate over tuples in a dataframe 
for row in data.itertuples(index=True, name='Pandas'):
    if getattr(row, "Disease") == "Septoria":
        print (getattr(row, "Remedy"))
'''   
    
