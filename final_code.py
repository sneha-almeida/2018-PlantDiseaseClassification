import tkinter
from tkinter import *
from PIL import Image
import os
from tkinter import filedialog
import numpy as np
import cv2
import skimage
from skimage import io
from skimage.io import imread_collection
#from PIL import image
import pandas as pd
import skimage.feature as sk
import numpy as np
import openpyxl
from sklearn import svm
from sklearn.svm import SVC 
from sklearn.model_selection import train_test_split
import openpyxl
from sklearn.preprocessing import normalize
from sklearn.neural_network import MLPClassifier  #Library for BPNN
from sklearn.metrics import classification_report, confusion_matrix
import xlrd
import ast
import pickle


root= Tk()

def resizeIt():
    filename = filedialog.askopenfilename()
    basewidth = 300
    img = Image.open(filename)#image is received here 
    print(img)
    wpercent = (basewidth / float(img.size[0]))
    hsize = int((float(img.size[1]) * float(wpercent)))
    img = img.resize((basewidth, hsize), Image.ANTIALIAS)
    img.save('resize.jpg')

    #perform the required operations on the image
    #io.imshow (images[i]) #display the images 
    #image = images[i]
    image = cv2.imread('resize.jpg')
    image = cv2.cvtColor(image,cv2.COLOR_BGR2GRAY)#convert the images to gray scale
    
    edges = cv2.Canny(image,100,200) #edge dection algorithm(canny edge detection) is used and a new image which just consists of the edge of original image is stored in edges
    
    img_final = edges - image #finally on the required pixels are taken and rest of the image is blurred out
    

    print("Values in numpy ")
    print(np.array(img_final)) #print the matrix which makes up the image

    #Extracting the features of an image using greycomatrix and normalizing the resultant matrix
    print("Extracgting features of the image......")
    image = sk.greycomatrix(img_final,[1],[0], 256, symmetric=True, normed=True)

    #image normalization

    #print(image.shape)
    #image = image.reshape(256,256,1,1)

    #Using the greycoprops function
    #Contrast
    img_contrast = sk.greycoprops(image,'contrast')[0,0]
    a = np.array(img_contrast)
    contrast = np.asscalar(a)
    print("*******************Extracting Feature---------Contrast******************")
    print(img_contrast)
    #Dissimilarity
    img_dissimilarity = sk.greycoprops(image,'dissimilarity')[0,0]
    a = np.array(img_dissimilarity)
    dissimilarity = np.asscalar(a)
    print("*******************Extracting Feature---------dissimilarity******************")
    print(img_dissimilarity)
    #Energy
    img_energy = sk.greycoprops(image,'energy')[0,0]
    a = np.array(img_energy)
    energy = np.asscalar(a)
    print("*******************Extracting Feature---------energy******************")
    print(img_energy)
    #ASM
    img_ASM = sk.greycoprops(image,'ASM')[0,0]
    a = np.array(img_ASM)
    asm = np.asscalar(a)
    print("*******************Extracting Feature---------ASM******************")
    print(img_ASM)
    #Correlation
    img_correlation = sk.greycoprops(image,'correlation')[0,0]
    a = np.array(img_correlation)
    correlation = np.asscalar(a)
    print("*******************Extracting Feature---------correlation******************")
    print(img_correlation)
    #Homogenity
    img_homogeneity = sk.greycoprops(image,'homogeneity')[0,0]
    a = np.array(img_homogeneity)
    homogeneity = np.asscalar(a)
    print("*******************Extracting Feature---------homogeneity******************")
    print(img_homogeneity)
    
    img_matrix = []
    img_matrix = np.array(image) #store it in an array
    print("Values in array........................")
    print(img_matrix)# print that array which is actually the original matrix in a new data container

    
    #creates a file in the directory/folder where the .py file resides and save the matrix to a .txt file
    #np.savetxt('file.txt',(img_matrix),header='Start of the file', footer='End of the file',newline = '\n\n\n')
    #creates a file in the directory/folder where the .py file resides
    
    #file = open(filename+str(filenum)+".txt",'w')
    #img_matrix.tofile(file, format = "%s",sep = " ")
    #filenum = filenum + 1
    #print("************************************Written to file ",i+1,"**********************************************")

    print("Written to file") #acknowledgement printing

    #Finding the dimensions of a matrix
    print("Shape of the greycomatrix",image.shape)
    print("Shape of the contrast matrix",img_contrast.shape)
    print("Shape of the dissimilarity matrix",img_dissimilarity.shape)
    print("Shape of the ASM matrix",img_ASM.shape)
    print("Shape of the correlation matrix",img_correlation.shape)
    print("Shape of the homogeneity matrix",img_homogeneity.shape)

    #calculate the moments for the image
    m = cv2.moments(img_final)
    print("Moment is:",m)


    #calculate the variance of the image
    variance = np.var(img_final)
    print (variance)

    #calculate the mean of the image
    mean = np.mean(img_final)
    print (mean)

    #open the excel file to be updated
    filename = 'E://test_image.xlsx'
    wb = openpyxl.load_workbook(filename)

    #Gives the number of worksheet in the excel file
    for ws in wb.worksheets:
        print(ws.title)

    #create the data which has to be written to the excel file
    ws = wb.worksheets[0]
    c = [contrast,dissimilarity,asm,correlation,homogeneity,variance,mean,energy]

    #findt the last row where the data is to be appended
    #row = ws.max_row

    #iterates over the excel file to write the data to the excel file
    #here the row will be constant and column will be iterated
    
    #for j in range(len(c)):
        #for l in c.values.tolist():
            #for item in l:
                #ws.cell(row=row, column=j).value = item
  
    
    for j in range(len(c)):
        ws.cell(row=2, column=j+1).value = c[j]
    #r = r+1

    #Save the data written to the file
    wb.save("E://test_image.xlsx")
    
    #use fromfile() to reload data
    ##############################Put a comment here#############################
    
    #classifier for SVM
    
    print("**************************SVM*****************************************")
    filename = 'E:\plant_classification.xlsx'
    data = pd.read_excel(filename)

    z = data.drop('Class', axis=1)
    #x = z.drop('ASM', axis=1)
    y = data['Class']

    #print(x)
    #print(y)

    train_data,test_data,train_label,test_label =train_test_split(z,y,test_size = 0.1)

    #print(train_data)
    #print(train_label)
    #print(test_data)
    #print(test_label)

    # training a linear SVM classifier 

    '''
    print("Creating Classifier..............")
    clf = SVC(kernel = 'linear', C = 1)

    print("Training..................")
    clf.fit(train_data,train_label)
    print("Training Complete................")
    filename = 'E://finalized_model_svm.sav'
    pickle.dump(model, open(filename, 'wb'))
    '''
    filename = 'E://finalized_model_svm.sav'
    #load the model
    loaded_model = pickle.load(open(filename, 'rb'))
    #testing
    predict = loaded_model.predict(test_data)
    print(predict)

    #Accuracy testing
    acc = np.mean(predict==test_label)
    print("Accuracy = ",acc*100)

    #metrics
    print (confusion_matrix(test_label,predict))
    print (classification_report(test_label,predict))

    test_image_svm = [(round(img_contrast,2),round(img_dissimilarity,2),round(img_correlation,2),round(img_homogeneity,2),round(variance,2),round(mean,2),round(energy,2))]
    #dict = ast.literal_eval(test_image) 
    #print(type(test_image))
    #df_svm = pd.DataFrame(test_image, columns =['contrast', 'dissimilarity','correlation','homogeneity','variance','mean','energy'],index=None)
    
    
    #print(test_image.shape)
    #file = 'E://test_image.xlsx'
    #test_image = pd.read_excel(file)
    test_predict_svm = loaded_model.predict(test_image_svm)
    print(test_predict_svm)
    window = Tk()
    l = Label(window, text="Disease Predicted by SVM")
    l.grid(row = 2, column = 0)

    l2 = Label(window, text=test_predict_svm)
    l2.grid(row = 2, column = 4)

    l2 = Label(window, text="        ")
    l2.grid(row = 2, column = 5)


    
    #BPNN Classifier
    print("****************************************BPNN*******************************")
    filename = 'E:\plant_classification.xlsx'
    data = pd.read_excel(filename)

    z = data.drop('class', axis=1)
    y = data['class']

    #print(x)
    #print(y)

    train_data,test_data,train_label,test_label =train_test_split(z,y,test_size = 0.1)
    '''
    print(type(test_data))
    print(test_data.shape)

    print("Creating Classifier.........")
    mlp = MLPClassifier(hidden_layer_sizes=(10, 10, 10), max_iter=100)
    print("training the classifier")
    mlp.fit(train_data, train_label)
    print("Training Complete")
    '''
    filename='E://finalized_model_bpnn.sav'
    #load the model
    loaded_model_svm = pickle.load(open(filename, 'rb'))
    #predictions
    predictions = loaded_model_svm.predict(test_data)
    print("******Predicted Data********")
    print(predictions)

    #Accuracy Calculations
    accuracy = np.mean(predictions==test_label)
    print("Accuracy = ",accuracy*100)  

    #metrics
    print (confusion_matrix(test_label,predictions))
    print (classification_report(test_label,predictions))
    #test_image = [[],[],[],[],[],[],[]]
    test_image = [(round(img_contrast,2),round(img_dissimilarity,2),round(img_correlation,2),round(img_homogeneity,2),round(variance,2),round(mean,2),round(energy,2))]
    #dict = ast.literal_eval(test_image) 
    print(type(test_image))
    #df = pd.DataFrame(test_image, columns =['contrast', 'dissimilarity','correlation','homogeneity','variance','mean','energy'],index=None)
    
    
    #print(test_image.shape)
    #file = 'E://test_image.xlsx'
    #test_image = pd.read_excel(file)
    test_predict = loaded_model_svm.predict(test_image)
    print(test_predict)

    window1 = Tk()
    l4 = Label(window1, text="Clothes Identification")
    l4.grid(row = 2, column = 0)

    l6 = Label(window1, text=test_predict)
    l6.grid(row = 2, column = 4)

    l8 = Label(window1, text="        ")
    l8.grid(row = 2, column = 5)
    

Button(text='add image', command=resizeIt).pack()

root.mainloop()
